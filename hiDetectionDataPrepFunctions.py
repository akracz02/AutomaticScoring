import cv2
import numpy as np
import pandas as pd

from sklearn.neighbors import KNeighborsClassifier
from openpyxl import load_workbook
from typing import Tuple

from imageProcessing import reduceImageOfEllipseAndGetNewCenter, reduceImageOfEllipse,\
getTransformationParameters, getTransformedImage, getBoundriesAndMask, reduceImageAndRemoveBackground,\
detectHit


def prepareDataSet() -> KNeighborsClassifier:
    knn = KNeighborsClassifier(n_neighbors=3)
    data = pd.read_excel('hitDetectionData.xlsx', sheet_name='Data')
    noArrStd = data['noHitStd'].to_numpy()
    noArrMax = data['noHitMax'].to_numpy()
    arrStd = data['hitStd'].to_numpy()
    arrMax = data['hitMax'].to_numpy()

    std_train = np.concatenate([noArrStd, arrStd], axis=0)
    max_train = np.concatenate([noArrMax, arrMax], axis=0)
    y_train = np.concatenate([np.array([0 for _ in range(len(noArrStd))]), np.array([1 for _ in range(len(arrStd))])])


    knn.fit(np.array([std_train, max_train]).T, y_train)
    return knn


def addToDataSet(ppframe: np.ndarray, pframe: np.ndarray, frame: np.ndarray):
    framecpy = cv2.cvtColor(frame.copy(), cv2.COLOR_RGB2GRAY)
    pframecpy = cv2.cvtColor(pframe.copy(), cv2.COLOR_RGB2GRAY)
    ppframecpy = cv2.cvtColor(ppframe.copy(), cv2.COLOR_RGB2GRAY)

    frame_pframe_diff = np.abs((framecpy.astype(np.int16) - pframecpy.astype(np.int16))).astype(np.uint8)
    pframe_ppframe_diff = np.abs((pframecpy.astype(np.int16)-ppframecpy.astype(np.int16))).astype(np.uint8)

    fpf_std = np.std(frame_pframe_diff)
    pfppf_std = np.std(pframe_ppframe_diff)

    fpf_max = np.max(frame_pframe_diff)
    pfppf_max = np.max(pframe_ppframe_diff)

    fpf_sumChanged = np.sum(frame_pframe_diff > np.mean(frame_pframe_diff))
    pfppf_sumChanged = np.sum(pframe_ppframe_diff > np.mean(pframe_ppframe_diff))

    workbook = load_workbook('hitDetectionData.xlsx')
    dataSheet = workbook['Data']

    noArrowHistValsSheet = workbook['noArrowHistogramValues']
    arrowHistValsSheet = workbook['arrowHistogramValues']

    lastrow = 0
    for row in range(1, dataSheet.max_row+2):
        if dataSheet.cell(row=row, column=1).value is not None:
            lastrow = row
        else:
            dataSheet.cell(row=lastrow+1, column=1, value=pfppf_std)
            dataSheet.cell(row=lastrow+1, column=2, value=pfppf_max)
            dataSheet.cell(row=lastrow+1, column=3, value=pfppf_sumChanged)

            dataSheet.cell(row=lastrow+1, column=4, value=fpf_std)
            dataSheet.cell(row=lastrow+1, column=5, value=fpf_max)
            dataSheet.cell(row=lastrow+1, column=6, value=fpf_sumChanged)

    noArrowHist, _ = np.histogram(pframe_ppframe_diff.ravel(), bins=256, range=[0,256])
    arrowHist, _ = np.histogram(frame_pframe_diff.ravel(), bins=256, range=[0,256])
    lastcol = 0
    for col in range(1, noArrowHistValsSheet.max_column+2):
        if noArrowHistValsSheet.cell(row=1, column=col).value is not None:
            lastcol = col
        else:
            for row in range(0,256):
                noArrowHistValsSheet.cell(row=row+1,column=lastcol+1,value=noArrowHist[row])
                arrowHistValsSheet.cell(row=row+1,column=lastcol+1,value=arrowHist[row])
    workbook.save('hitDetectionData.xlsx')
    return


def getEllipse(frame: np.ndarray) -> Tuple[Tuple[float,float], Tuple[float,float], float]:
    # Initialize a list to store points
    global points
    points = []

    # Mouse callback function to capture the points
    def select_point(event, x, y, flags, param):
        if event == cv2.EVENT_LBUTTONDOWN:  # Check for left mouse button click
            points.append((x, y))
            print(f"Point selected: {x}, {y}")
            # Draw a small circle where the user clicked
            cv2.circle(image, (x, y), 5, (0, 0, 255), -1)  # Red dot
            cv2.imshow("Image", image)
            # Stop after 5 points
            if len(points) == 5:
                print("Selected points:", points)
                cv2.destroyAllWindows()

    # Load the image
    #image_path = "your_image.jpg"
    #image = cv2.imread(image_path)
    image = frame.copy()
    # Create a window and set the mouse callback function
    cv2.imshow("Image", image)
    cv2.setMouseCallback("Image", select_point)

    # Keep the window open until the user closes it
    cv2.waitKey(0)

    # Optional: Print and save the selected points
    print("Final selected points:", points)
    return cv2.fitEllipse(np.array(points))


def getFrames(video_path: str, frame_idx = 0) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    i = 0
    # Create a VideoCapture object
    cap = cv2.VideoCapture(video_path)

    pframe = None
    ppframe = None
    frame = None
    rawframe = None

    # Check if the video file opened successfully
    if not cap.isOpened():
        print("Error: Cannot open video file")
    else:
        while True:
            if i%4 == 0:
                ppframe = pframe
                pframe = frame
                frame = rawframe

                # Read a frame from the video
            ret, rawframe = cap.read()

            # If no frame is returned, end of video
            if not ret:
                print("End of video")
                break
                
                # Display the frame
            cv2.imshow("Frame", rawframe)

            # Exit if 'q' is pressed
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

            i+=1

            if i == frame_idx:
                break
            

        # Release the VideoCapture object
        cap.release()
        cv2.destroyAllWindows()

    return frame, pframe, ppframe


def testDetection(video_path: str, ell: Tuple[Tuple[float,float],Tuple[float,float],float], knn: KNeighborsClassifier):
    i = 0
    # Create a VideoCapture object
    cap = cv2.VideoCapture(video_path)

    pframe = None
    ppframe = None
    frame = None
    rawframe = None

    newell = None
    imred = None
    imtrans = None
    rotMat = None
    scMat = None
    center = None
    bnds = None
    mask = None


    # Check if the video file opened successfully
    if not cap.isOpened():
        print("Error: Cannot open video file")
    else:
        while True:
            # Read a frame from the video
            ret, rawframe = cap.read()

            if i%100 == 0:
                ppframe = pframe
                pframe = frame
                frame = rawframe
                if ppframe is not None and pframe is not None and frame is not None:
                    if newell is None:
                        imred, newell = reduceImageOfEllipseAndGetNewCenter(frame, ell)
                    else:
                        imred = reduceImageOfEllipse(frame, ell)
                    pimred = reduceImageOfEllipse(pframe, ell)
                    ppimred = reduceImageOfEllipse(ppframe, ell)

                    if rotMat is None:
                        imtrans, rotMat, scMat, center = getTransformationParameters(imred, newell)
                    else:
                        imtrans = getTransformedImage(imred, rotMat, scMat)
                    pimtrans = getTransformedImage(pimred, rotMat, scMat)
                    ppimtrans = getTransformedImage(ppimred, rotMat, scMat)

                    if bnds is None:
                        bnds, mask = getBoundriesAndMask(imtrans, center, newell[1][1])

                    imFinal = reduceImageAndRemoveBackground(imtrans, bnds, mask)
                    pimFinal = reduceImageAndRemoveBackground(pimtrans, bnds, mask)
                    ppimFinal = reduceImageAndRemoveBackground(ppimtrans, bnds, mask)

                    if detectHit(ppimFinal, pimFinal, imFinal, knn):
                        print('Hit!')

            # If no frame is returned, end of video
            if not ret:
                print("End of video")
                break
                
                # Display the frame
            cv2.imshow("Frame", rawframe)

            # Exit if 'q' is pressed
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

            i+=1
            

        # Release the VideoCapture object
        cap.release()
        cv2.destroyAllWindows()